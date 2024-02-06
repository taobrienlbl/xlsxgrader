import pandas as pd
from pathlib import Path
import openpyxl

def parse_canvas_csv(csv_file_path : str | Path):
    """ Parse a CSV file exported from Canvas and return a cleaned-up DataFrame. """
    # load the CSV file; index is student name
    data = pd.read_csv(csv_file_path, index_col='name')

    # determine the number of questions
    # the columns are formatted as follows: id, section, section_id, submitted, attempt, question1, score1, question2, score2, question3, score3, ..., n correct, n incorrect, score
    # we can use the number of columns to determine the number of questions
    informational_columns_start = ['id', 'section', 'section_id', 'submitted', 'attempt']
    informational_columns_end = ['n correct', 'n incorrect', 'score']
    informational_columns = informational_columns_start + informational_columns_end
    num_informational_columns = len(informational_columns)
    num_informational_columns_start = len(informational_columns_start)
    num_informational_columns_end = len(informational_columns_end)
    num_columns = len(data.columns)
    num_questions = (num_columns - num_informational_columns) // 2

    # check that the number of columns is consistent with the number of questions
    if num_columns != num_informational_columns + 2 * num_questions:
        raise RuntimeError(f"Unexpected number of columns: {num_columns}")

    # extract the response columns
    response_columns = data.columns[num_informational_columns_start:-num_informational_columns_end:2]

    # extract the grade columns
    grade_columns = data.columns[num_informational_columns_start+1:-num_informational_columns_end:2]

    # extract the questions and the answers; copy the data to a new dataframe
    responses = data[response_columns].copy()
    grades = data[grade_columns].copy()

    # determine which questions need to be graded by checking if all the scores are 0
    grade_columns_that_sum_to_zero = data[grade_columns].sum() == 0

    # Extract the question text from each response column name
    question_text = [column.split(': ')[1] for column in responses.columns]

    # Extract the maximum score for each question from the grade column names
    # note that Pandas automatically appends a .1, .2, etc. to the column names if there are duplicates
    # so we need to extract the unique question numbers and then find the corresponding maximum score
    grade_column_names = grades.columns
    # check if there is more than one decimal point in the column names, and remove everything following the second decimal point
    maximum_grades = [int(column.split('.', 2)[0]) for column in grade_column_names]

    # generate new column names as Question 1, Question 2, etc.
    new_column_names = [f'Question {i+1}' for i in range(len(question_text))]

    # rename the columns for responses and grades
    responses.columns = new_column_names
    grades.columns = new_column_names

    # get the columns of the questions to grade (this is done after renaming the columns to Question 1, Question 2, etc.)
    responses_to_grade_bool = grade_columns_that_sum_to_zero.values
    responses_to_grade_columns = responses.columns[responses_to_grade_bool]

    # make a pandas dataframe for the maximum grades
    max_grades_df = pd.DataFrame(maximum_grades, index=new_column_names, columns=['Max Grade']).T

    # make a pandas dataframe for the question text
    question_text_df = pd.DataFrame(question_text, index=new_column_names, columns=['Question Text']).T

    # parse the last name and first name from the index
    last_names = [name.split(' ')[-1] for name in data.index]
    first_names = [' '.join(name.split(' ')[:-1]) for name in data.index]

    # add last name and first name columns to grades and responses
    grades.loc[:,'Last Name'] = last_names
    grades.loc[:,'First Name'] = first_names
    responses.loc[:,'Last Name'] = last_names
    responses.loc[:,'First Name'] = first_names

    # make last name and first name the first two columns
    grades = grades[['Last Name', 'First Name'] + list(grades.columns[:-2])]
    responses['Last Name'] = last_names
    
    # sort the dataframes by last name
    grades = grades.sort_values(by='Last Name')
    responses = responses.sort_values(by='Last Name')

    # collect the question data into a dictionary
    question_data = {
        'question_text': question_text_df,
        'responses': responses,
        'grades': grades,
        'responses_to_grade': responses_to_grade_columns,
        'max_grades': max_grades_df
    }

    return question_data

def save_to_xlsx(question_data, xlsx_file_path : str | Path):
    """ Save the question data to an xlsx file. """

    # set the worksheet names
    worksheet_names = ["Total Scores"] + list(question_data['responses'].columns)

    # remove Last Name and First Name from the list of worksheet names
    worksheet_names.remove('Last Name')
    worksheet_names.remove('First Name')

    # create a new workbook
    wb = openpyxl.Workbook()

    # rename the first sheet
    wb.active.title = worksheet_names[0]

    # create the worksheets
    for sheet_name in worksheet_names[1:]:
        wb.create_sheet(sheet_name)

    # insert the question/response data into the question worksheets
    for sheet_name in worksheet_names[1:]:
        ws = wb[sheet_name]
        # set cell C1 to the question text for each question worksheet
        ws['C1'] = question_data['question_text'][sheet_name].values[0]
        ws.append(['Student Name', 'Response', "Score", "Max Score", "Comments"])
        # make the first row italic
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(italic=True)
        # make the second row bold
        for cell in ws[2]:
            cell.font = openpyxl.styles.Font(bold=True)
        # iterate over the rows in the responses dataframe and insert them into the worksheet
        for row, grade_row in zip(question_data['responses'].iterrows(), question_data['grades'].iterrows()):
            # name
            new_row = [row[0]]
            # response
            new_row.append(row[1][sheet_name])
            # score
            new_row.append(grade_row[1][sheet_name])
            # max score
            new_row.append(question_data['max_grades'][sheet_name].values[0])
            # comments
            new_row.append('')
            ws.append(new_row)

        # set the width of the second column and turn on text wrapping
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['B'].bestFit = True
        ws.column_dimensions['B'].auto_size = True

        # set row heights based on column B
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

        # freeze rows 1 and 2
        ws.freeze_panes = ws['A3']

        # hide column A
        ws.column_dimensions['A'].hidden = True


    # create the total scores worksheet from scratch
    ws1 = wb[worksheet_names[0]]
    ws1.append([]) # add an empty row so that this sheet matches the others
    columns = ['Full Student Name', 'Last Name', 'First Name', 'Total Score', 'Comments']
    ws1.append(columns)
    # make the first row bold
    for cell in ws1[2]:
        cell.font = openpyxl.styles.Font(bold=True)
    # set the current row
    start_row = 3

    current_color = "FFFFFF"

    # populate the rows
    for n, row in enumerate(question_data['responses'].iterrows()):
        # set the current row
        current_row = start_row + n

        # name
        new_row = [row[0]]

        # last name
        new_row.append(row[1]['Last Name'])
        
        # first name
        new_row.append(row[1]['First Name'])

        # total score (use formulas to connect to the data in the other sheets)
        total_score_formula = "=SUM("
        for question in worksheet_names[1:]:
            total_score_formula += f"'{question}'!C{current_row},"
        total_score_formula = total_score_formula[:-1] + ")"
        # total score
        new_row.append(total_score_formula)

        # comments
        # use a formula to construct the comments in the following format:
        # "Question 1 (score/max score): comments; Question 2 (score/max score): comments; ..."
        comment_formula = "=CONCATENATE("

        for n, question in enumerate(question_data['responses_to_grade']):
            item = f'"{question} (",'
            item += f"'{question}'!C{current_row},"
            item += '"/",'
            item += f"'{question}'!D{current_row}"
            item += ',") ",'
            item += f"'{question}'!E{current_row}"
            if n < len(question_data['responses_to_grade']) - 1:
                item += ',CHAR(10),"|",CHAR(10),'
            else:
                item += ')'
            comment_formula += item
        new_row.append(comment_formula)

        # append the row
        ws1.append(new_row)

        # set the color of the row
        for cell in ws1[current_row]:
            cell.fill = openpyxl.styles.PatternFill(start_color=current_color, end_color=current_color, fill_type="solid")

        # alternate the color
        if current_color == "FFFFFF":
            current_color = "DDDDDD"
        elif current_color == "DDDDDD":
            current_color = "FFFFFF"
        
        # increment the current row
        current_row += 1
    
    # make the student name column bold
    for cell in ws1['A']:
        cell.font = openpyxl.styles.Font(bold=True)
    
    # auto set the name column width
    ws1.column_dimensions['A'].width = 15

    # set the width of the grade column
    ws1.column_dimensions['D'].width = 15

    # set the width of the comment column
    ws1.column_dimensions['E'].width = 50

    for cell in ws1['D']:
        # use center justification for the total score column
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
        # format the total score column to round to the first decimal place
        cell.number_format = '0.0'

    # put the maximum score in cell D1
    ws1['D1'] = f"Max Score: {sum(question_data['max_grades'].values[0])}"

    # freeze column A
    ws1.freeze_panes = ws1['B1']

    # hide columns B and C
    ws1.column_dimensions['B'].hidden = True
    ws1.column_dimensions['C'].hidden = True

    # get the first question to be graded
    first_question_to_grade = str(question_data['responses_to_grade'][0])

    # set the active worksheet to be the first question that needs to be graded
    wb.active = wb[first_question_to_grade]

    # save the xlsx file
    wb.save(xlsx_file_path)

    return
